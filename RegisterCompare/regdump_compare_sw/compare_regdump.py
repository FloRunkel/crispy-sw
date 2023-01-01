from abc import abstractmethod 
import sys
import argparse 
import openpyxl
from openpyxl.utils import get_column_letter
from typing import Dict, List
from pathlib import Path 
import os

nonusedtypes    =   ['raz']
nocomparenames  =   ['caseqsram', 'dqseqsram']
regtypes        =   {'ampsdq' :'LP5DDRPHY_AMPS_DQ', 'amphdq' : 'LP5DDRPHY_DQ' , 'impcal' : 'LP5DDRPHY_IMPCAL', 'amphca' : 'LP5DDRPHY_CA', 'ampsca' : 'LP5DDRPHY_AMPS_CA', 'ddrpll' : 'LP5DDRPHY_DDRPLL' }
    
 
def debug_excel_write(filename, printlist, header, sheetname='data', addsheet=False, nolist=False):
    if addsheet:
        wb = openpyxl.load_workbook(filename)
        ws_write = wb.create_sheet(sheetname) 
    else:
        wb = openpyxl.Workbook()
        ws_write = wb['Sheet']
        ws_write.title = sheetname  
    for i in range(len(header)):
        ws_write.cell(row=1,column=i+1).value = header[i]
    
    cnt = 2
    for i in printlist: 
        # print (i)
        for j in range(len(i)):
            cell = ws_write.cell(row=cnt,column=j+1) 
            cell.value = i[j] if nolist else i.list[j]
        cnt += 1

    for column_cells in ws_write.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        ws_write.column_dimensions[get_column_letter(column_cells[0].column)].width = length

    wb.save(filename=filename)

def as_text(value):
    return "" if value is None else str(value)


def print_objlist(filename:str, objectlist:List) -> None:

    with open(filename, 'w') as f:  
        for i in objectlist: 
            print(i, file=f) 
            for z in i:
                print (z, file=f) 


def print_result_obj(filename:str, objectlist:List) -> None:
    with open(filename, 'w') as f:  
        for i in objectlist:
            print(i, file=f)  


def replace_brackets(name:str) -> str:
    name = name.replace('[', '')
    name = name.replace(']', '') 
    return name

'''  
base classes
with those classes register and fields can be defined. This makes it easy to compare the objects
In order to no retype every function a base class is used to implement a minimum of required functions in every child class

'''
class BaseClass:

    def __init__(self, name:str) -> None:
        self._name = name 

    @property
    def name(self) -> str:
        return self._name

    @name.setter
    def name(self,new_name:str) -> None:
        self._name = new_name

    @abstractmethod
    def list(self) -> List:
        pass

    def __str__(self) -> List:
        return f'{self.list}'

    def __iter__(self):
        self._cnt = 0
        return self

    def __next__(self):
        if self._cnt < len(self.list):
            listresult = self.list[self._cnt] 
            self._cnt += 1
            return listresult
        else:
            raise StopIteration


    def __len__(self):
        return len(self.list)

    


class Field(BaseClass):

    def __init__(self, fieldname:str, fieldmsb:int, fieldlsb:int, fieldtype:str, fieldvalue:int) -> None:

        super().__init__(name=fieldname) 
        self._fieldmsb = fieldmsb 
        self._fieldlsb = fieldlsb 
        self._fieldtype = fieldtype
        self._fieldvalue = int(fieldvalue,base=16)
            
 
    @property
    def bits(self) -> str:
        bitstr = f'{self._fieldmsb}:{self._fieldlsb}'
        return bitstr

    @property
    def value(self) -> int:
        return self._fieldvalue

    @property
    def type(self) -> str:
        return self._fieldtype
    
    @property
    def list(self) -> List:
        namelist = [ self._name, self._fieldtype, self._fieldvalue] 
        return namelist
 


class Register(BaseClass):

    def __init__(self, registername:str, fieldlist:List) -> None:
        super().__init__(name=registername)  
        self._fieldlist = fieldlist
        self._fieldobject = [] 
        for fields in fieldlist: 
           self._fieldobject.append(Field(fieldname=fields['name'], fieldmsb=fields['msb'], fieldlsb=fields['lsb'], fieldtype=fields['type'], fieldvalue=hex(fields['value'])))
            
 
    @property
    def list(self) -> List:
        namelist = [self._name]
        for i in self._fieldobject:
            namelist.append(i.name)
        return namelist

    @property
    def fieldlist(self) -> List:
        return self._fieldobject
 

    def __next__(self):
        if self._cnt < len(self.fieldlist):
            listresult = self.fieldlist[self._cnt] 
            self._cnt += 1
            return listresult
        else:
            raise StopIteration

    def __len__(self):
        return len(self.fieldlist)


class Results(BaseClass):

    def __init__(self, registername:str, fieldname:str, fieldtype:str, refval:int, compval:int) -> None:
        super().__init__(name=registername)
        self._fieldname = fieldname
        self._fieldtype = fieldtype
        self._refval = refval 
        self._compval = compval 

    @property
    def regname(self) -> str:
        return self._name

    @property
    def fieldname(self) -> str:
        return self._fieldname
 
    @property
    def list(self) -> List:
        namelist = [self._name, self._fieldname, self._fieldtype, self._refval, self._compval]
        return namelist
        
 


class RegisterDataBase:

    def __init__(self,fname:str,regname:str, substring:str='') -> None:  
        self._fname = os.path.join(Path(__file__).parent,fname)
        self._regname = regname
        self._substring = substring
        self._regdatabase = []
 
    
    def read_txtfile(self) -> Dict: 
   
        """ parse textfile and get register + fields """
      
        file_dic = {} # this is a dictionary
        
        assert os.path.exists(self._fname), f"This file {self._fname} doesn't exist !"
        fh = open(self._fname,'r')
        lines = fh.readlines()
        fh.close()

        foundreg = 0
        for line in lines:
            line = line.strip()  # remove empty lines and whitespaces
            if line.strip():
                if (self._regname in line) and (self._substring in line):
                    hierpath = line.split('.')
                    for i in range(len(hierpath)):
                        if self._regname in hierpath[i]:
                            index = i
                            break
                    del hierpath[:index]
                    regsave = '.'.join(hierpath)
                    foundreg = 1
                    file_dic[regsave] = {}
                    fieldlist = []
                elif foundreg == 1:
                    words = line.split()
                    if '0x' not in words[0]: fieldlist.append(line)
                    if (':0]' in words[0]) or ('0]' in words[1]): 
                        foundreg = 0
                        file_dic[regsave] = fieldlist
        
        return file_dic

 
    def generate_database(self, registerdic:Dict):

        for registername in registerdic:
            fieldlist = []
            for fields in registerdic[registername]:
                fielddic = {}
                field = fields.split()  
                field = [replace_brackets(listfield) for listfield in field]
                if len(field) == 5:  
                    field = [listfield.replace(':', '') for listfield in field]
                    msb,lsb,value,regtype,name = field
                elif len(field) == 4:
                    fieldsplit = field[0].split(':')
                    msb,lsb = fieldsplit
                    value,regtype,name = field[1:]
                if regtype not in nonusedtypes:
                    fielddic['name'] = name.lower()
                    fielddic['msb'] = msb
                    fielddic['lsb'] = lsb
                    fielddic['type'] = regtype
                    fielddic['value'] = int(value,base=16)
                    fieldlist.append(fielddic)
            nosave = False
            for comparetest in nocomparenames: 
                if comparetest in registername: nosave = True
            if not nosave:
                regobj = Register(registername,fieldlist)
                self._regdatabase.append(regobj)

        return self._regdatabase


class CompareRegister:

    def __init__(self,referencelist:List, checklist:List) -> None:
        self._referencelist = referencelist
        self._checklist = checklist


    def compare_list_element(self,list1:List, list2:List) -> bool:   
        if len(list1) != len(list2):
            return False
        else:
            bool_list = []
            for i in range(len(list1)):
                bool_list.append(list1[i] == list2[i])
        return all(bool_list)

        #bool_list = list(map(lambda x, y: x == y, list1, list2))
        #return all(bool_list)

    def compare_cells (self, registername:str) -> None:

        resultobjlist = []
        noregs = []
        nofields = []
        for refregs in self._referencelist: 
            if registername in refregs.name:
                compare_regname =  refregs.name
                refregobj = refregs
                reflist = [compare_regname]
                foundreg = False
                for fields in refregs:
                    reflist.append(fields.name)
                    reflist.append(fields.value) 
                for compregs in self._checklist:
                    if compregs.name == compare_regname:
                        compregsobj = compregs
                        foundreg = True
                        complist = [compregs.name]
                        for fields in compregs:
                            complist.append(fields.name)
                            complist.append(fields.value)
                if not foundreg:
                    noregs.append([compare_regname])
                   # print (f'register {compare_regname} not found in checklist')
                else: 

                    if not self.compare_list_element(reflist,complist):
                        # print (f' reflist = {reflist} , complist = {complist}')
                       # print (f' registername = {compare_regname}')
 
                        for reffields in refregobj:
                            fieldfound = False
                            reffieldname = reffields.name
                            reffieldval = reffields.value
                            reffieldtype = reffields.type
                            for compfields in compregsobj:
                                if compfields.name == reffieldname:
                                    fieldfound = True
                                    if reffieldval != compfields.value:
                                        resultobj = Results(compare_regname,reffieldname, reffieldtype, reffieldval,compfields.value)
                                        resultobjlist.append(resultobj)
                               #         print (f'fieldname = {reffieldname} , reffieldvalue = {reffieldval} <=> compfieldvalue = {compfields.value}')
                            if not fieldfound:
                                nofields.append([compare_regname, reffieldname, reffieldtype])
                              #  print (f'this field {reffieldname} is not in the compare list')
        return resultobjlist, noregs , nofields



'''
main function, which is executed with calling the script.
The idea is to generate a Excel file with different sheets.
Each sheet contains the differences of the 2 files linked to the 
APB register hierarchy
'''


if __name__ == '__main__': 
    

    # Adding argparser for command line call
    if len(sys.argv) > 1:  

        argparser = argparse.ArgumentParser(description='This script process 2 different register dump files') 
        argparser.add_argument('-rf', '--ref_file',         action='store', help='path + filename to reference file')
        argparser.add_argument('-cf', '--comp_file',        action='store', help='path + filename to compare file')
        argparser.add_argument('-o',  '--out_file',         action='store', default='compare_result.xlsx', help='filename of output excel file')
        argparser.add_argument('-rn', '--regname',          action='store', default='LP5DDRPHY', help='Register name , which is used to parse')
        argparser.add_argument('-rs', '--ref_subname',      action='store', default='', help='subname just needed if there in more than 1 PHY listed') 
        argparser.add_argument('-cs', '--comp_subname',     action='store', default='GRP0', help='subname just needed if there in more than 1 PHY listed')
        argparser.add_argument('-dbg','--debug',            action='store_true', default=False, help='enable debug feature')
        args = argparser.parse_args()

        assert args.ref_file != None ,f'no path and filename given for reference file (-rf)'  
        assert args.comp_file != None ,f'no path and filename given for compare file file (-cf)'   
        referencefilename = args.ref_file
        comparefilename = args.comp_file 
        resultxls = args.out_file
        regname = args.regname 
        ref_subname = args.ref_subname 
        comp_subname = args.comp_subname 
        debug = args.debug
        debugdir = './'

    else:
        
        compare = 'warm'
        project = 'caicos'
        directory = f'{Path(__file__).parent}/../regdump_runs/{project}'
        inputdir = f'{directory}/project_inputs'
        debugdir = f'{directory}/debug'
        resultdir = f'{directory}/results'

        compare_select = {
            'cold' : {'secondfile' : f'{inputdir}/dcs0_amph_cold_phase31.txt' , 'resultxls' : f'{resultdir}/coldboot_compare.xlsx'} ,
            'warm' : {'secondfile' : f'{inputdir}/dcs0_amph_warm_phase31.txt' , 'resultxls' : f'{resultdir}/warmboot_compare.xlsx'} 
        }

        reference_select = {
            'cold' : f'{inputdir}/dcs_amph_cold_phase2.txt',
            'warm' : f'{inputdir}/dcs_amph_warm_phase2.txt'
        }
   
        referencefilename = reference_select[compare]
        comparefilename   = compare_select[compare]['secondfile']
        resultxls         = compare_select[compare]['resultxls'] 

        regname = 'LP5DDRPHY'
        ref_subname = ''
        comp_subname = ''

        debug = True


    #############################################
    # generate databases
    # ###########################################   


    # 1st list
    referenceobj = RegisterDataBase(referencefilename,regname,ref_subname)
    referencedic = referenceobj.read_txtfile()
    referenceobjlist = referenceobj.generate_database(referencedic)
 
    if debug: print_objlist(f'{debugdir}/referenceobj.txt', referenceobjlist) 
             
 
    # 2nd list 
    compareobj = RegisterDataBase(comparefilename,regname,comp_subname)
    comparedic = compareobj.read_txtfile()
    compareobjlist = compareobj.generate_database(comparedic)
 
    if debug:  print_objlist(f'{debugdir}/compareobj.txt',compareobjlist) 

    # compare / should be linked to a config with a loop
    compareobj = CompareRegister(referenceobjlist,compareobjlist)
    resultslistall = {}
    noregslist = []
    nofieldslist = []
    for regskey in regtypes:
        resultlist, noregs, nofields = compareobj.compare_cells(regtypes[regskey])
        resultslistall[regskey] = resultlist
        noregslist = noregslist + noregs
        nofieldslist = nofieldslist + nofields


    if debug: print_result_obj(f'{debugdir}/listdq.txt',resultslistall['amphdq']) 


    header = ['registername','fieldname','fieldtype',f'{referencefilename}',f'{comparefilename}']
 
    addsheet = False
    for regskey in regtypes: 
        debug_excel_write(resultxls, resultslistall[regskey], header, sheetname=regskey, addsheet=addsheet)
        addsheet = True


    header = ['registername']
    debug_excel_write(resultxls, noregslist, header,sheetname='noregslist', addsheet=True, nolist=True)
    header = ['registername', 'fieldname', 'fieldtype']
    debug_excel_write(resultxls, nofieldslist, header,sheetname='nofieldlist', addsheet=True, nolist=True)
